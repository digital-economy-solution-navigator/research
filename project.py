import requests
import json
import pandas as pd
import re
from openpyxl.utils import get_column_letter

def get_manager_projects(manager_id):
    """
    Fetch all projects under a manager from the UNIDO Compass API.
    
    Args:
        manager_id (int): The manager ID to fetch projects for
    
    Returns:
        dict: JSON response containing the projects
    """
    url = f"https://compass.unido.org/api/v1/managers/{manager_id}/projects"
    
    try:
        response = requests.get(url)
        response.raise_for_status()  # Raise an exception for bad status codes
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data: {e}")
        return None

def get_project_details(project_id):
    """
    Fetch detailed information for a specific project from the UNIDO Compass API.
    
    Args:
        project_id (int): The project ID to fetch details for
    
    Returns:
        dict: Project details including focus_area, description, donors_json, supplier_json, partners_json
    """
    url = f"https://compass.unido.org/api/v1/projects/{project_id}"
    
    try:
        response = requests.get(url)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error fetching project {project_id} details: {e}")
        return None

def sanitize_sheet_name(name):
    """
    Sanitize sheet name to comply with Excel restrictions.
    Excel sheet names cannot contain: / \ ? * [ ]
    Maximum length is 31 characters.
    """
    # Remove invalid characters
    name = re.sub(r'[/\\?*\[\]]', '', name)
    # Truncate to 31 characters if needed
    if len(name) > 31:
        name = name[:31]
    return name

def extract_names_from_json(json_data, field_name='name'):
    """
    Extract names from a JSON array of objects.
    
    Args:
        json_data: JSON array (list) or JSON string
        field_name: The field to extract from each object (default: 'name')
    
    Returns:
        str: Comma-separated list of names, or empty string if no data
    """
    if not json_data:
        return ''
    
    # If it's already a list, use it directly
    if isinstance(json_data, list):
        data_list = json_data
    # If it's a string, try to parse it
    elif isinstance(json_data, str):
        try:
            data_list = json.loads(json_data)
        except (json.JSONDecodeError, TypeError):
            return ''
    else:
        return ''
    
    # Extract names from each object
    names = []
    for item in data_list:
        if isinstance(item, dict) and field_name in item:
            name = item[field_name]
            if name:  # Only add non-empty names
                names.append(str(name))
    
    return ', '.join(names) if names else ''

def extract_donor_names(donors_json):
    """Extract donor names from donors_json."""
    return extract_names_from_json(donors_json, 'name')

def extract_supplier_names(supplier_json):
    """Extract supplier names from supplier_json."""
    return extract_names_from_json(supplier_json, 'name')

def extract_partner_names(partners_json):
    """Extract partner names from partners_json."""
    return extract_names_from_json(partners_json, 'name')

def extract_country_names(all_countries_json):
    """
    Extract country names from all_countries_json.
    
    Handles JSON string format like: [{"id": "M2", "name": "Africa"}, {"id": "M0", "name": "Global"}]
    Returns comma-separated names like: "Africa, Global"
    """
    return extract_names_from_json(all_countries_json, 'name')

def count_countries(all_countries_json):
    """Count countries from all_countries_json."""
    return count_items_in_json(all_countries_json)

def count_items_in_json(json_data):
    """
    Count items in a JSON array.
    
    Args:
        json_data: JSON array (list) or JSON string
    
    Returns:
        int: Number of items in the array, or 0 if no data
    """
    if not json_data:
        return 0
    
    # If it's already a list, use it directly
    if isinstance(json_data, list):
        return len(json_data)
    # If it's a string, try to parse it
    elif isinstance(json_data, str):
        try:
            data_list = json.loads(json_data)
            return len(data_list) if isinstance(data_list, list) else 0
        except (json.JSONDecodeError, TypeError):
            return 0
    else:
        return 0

def count_donors(donors_json):
    """Count donors from donors_json."""
    return count_items_in_json(donors_json)

def count_suppliers(supplier_json):
    """Count suppliers from supplier_json."""
    return count_items_in_json(supplier_json)

def process_manager_projects(manager_id):
    """
    Process all projects for a given manager ID.
    Fetches projects and their detailed information.
    
    Args:
        manager_id (int): The manager ID to process
    
    Returns:
        tuple: (DataFrame with projects, manager_name, first_name) or (None, None, None) if failed
    """
    print(f"\nProcessing manager ID: {manager_id}")
    print("-" * 50)
    
    response = get_manager_projects(manager_id)
    
    if not response or "body" not in response:
        print(f"  ✗ Failed to fetch projects for manager {manager_id}")
        return None, None, None
    
    projects_data = response["body"]["data"]
    meta = response["body"]["meta"]
    
    if not projects_data:
        print(f"  ⚠ No projects found for manager {manager_id}")
        return None, None, None
    
    manager_name = f"{projects_data[0]['first_name']} {projects_data[0]['last_name']}"
    total_projects = meta.get('count', len(projects_data))
    
    print(f"  Manager: {manager_name}")
    print(f"  Total projects: {total_projects}")
    
    # Fetch additional details for each project
    print("  Fetching additional project details...")
    enhanced_projects = []
    for idx, project in enumerate(projects_data, 1):
        proj_id = project.get('proj_id')
        print(f"    [{idx}/{len(projects_data)}] Project {proj_id}...", end='\r')
        
        project_details = get_project_details(proj_id)
        
        # Create enhanced project record
        enhanced_project = project.copy()
        
        if project_details and "body" in project_details:
            body = project_details["body"]
            # The data is a list with one element containing the project details
            if "data" in body and isinstance(body["data"], list) and len(body["data"]) > 0:
                details = body["data"][0]
                # Extract the requested fields
                enhanced_project['focus_area'] = details.get('focus_area', '') or ''
                enhanced_project['description'] = details.get('description', '') or ''
                # Extract readable names from JSON arrays
                donors_data = details.get('donors_json', [])
                supplier_data = details.get('supplier_json', [])
                enhanced_project['donors_json'] = extract_donor_names(donors_data)
                enhanced_project['donor_count'] = count_donors(donors_data)
                enhanced_project['supplier_json'] = extract_supplier_names(supplier_data)
                enhanced_project['supplier_count'] = count_suppliers(supplier_data)
                enhanced_project['partners_json'] = extract_partner_names(details.get('partners_json', []))
                countries_data = details.get('all_countries_json', [])
                enhanced_project['all_countries_json'] = extract_country_names(countries_data)
                enhanced_project['recipient_country_count'] = count_countries(countries_data)
            else:
                enhanced_project['focus_area'] = ''
                enhanced_project['description'] = ''
                enhanced_project['donors_json'] = ''
                enhanced_project['donor_count'] = 0
                enhanced_project['supplier_json'] = ''
                enhanced_project['supplier_count'] = 0
                enhanced_project['partners_json'] = ''
                enhanced_project['all_countries_json'] = ''
                enhanced_project['recipient_country_count'] = 0
        else:
            enhanced_project['focus_area'] = ''
            enhanced_project['description'] = ''
            enhanced_project['donors_json'] = ''
            enhanced_project['donor_count'] = 0
            enhanced_project['supplier_json'] = ''
            enhanced_project['supplier_count'] = 0
            enhanced_project['partners_json'] = ''
            enhanced_project['all_countries_json'] = ''
            enhanced_project['recipient_country_count'] = 0
        
        enhanced_projects.append(enhanced_project)
    
    print(f"\n  ✓ Completed fetching details for {len(enhanced_projects)} projects")
    
    # Convert to DataFrame
    df = pd.DataFrame(enhanced_projects)
    
    # Remove manager columns (id, first_name, last_name) and proj_country_id first
    columns_to_drop = ['id', 'first_name', 'last_name', 'proj_country_id']
    for col in columns_to_drop:
        if col in df.columns:
            df = df.drop(columns=[col])
    
    # Rename columns: remove _json suffix and proj_ prefix
    column_rename_map = {
        'donors_json': 'donors',
        'supplier_json': 'supplier',
        'partners_json': 'partners',
        'all_countries_json': 'recipient_country_m',
        'proj_country_name': 'recipient_country',
        'proj_id': 'id',
        'proj_name': 'name',
        'proj_is_ongoing': 'is_ongoing',
        'proj_start_date': 'start_date',
        'proj_end_date': 'end_date',
        'proj_budget': 'budget',
        'proj_expenditure': 'expenditure',
        'proj_net_approval': 'net_approval'
    }
    df = df.rename(columns=column_rename_map)
    
    # Reorder columns to put project info first
    column_order = [
        'id', 'name', 'description', 'recipient_country_m', 'recipient_country_count', 'donors', 'donor_count',
        'is_ongoing', 'start_date', 'end_date',
        'budget', 'expenditure', 'net_approval',
        'focus_area', 'supplier', 'supplier_count', 'partners'
    ]
    
    # Only include columns that exist in the dataframe
    column_order = [col for col in column_order if col in df.columns]
    df = df[column_order]
    
    # Extract first name for later use
    first_name = manager_name.split()[0] if manager_name else f"Manager_{manager_id}"
    
    return df, manager_name, first_name

def generate_country_statistics(all_projects_df):
    """
    Generate statistics of project counts by recipient_country_m.
    When recipient_country_count > 1, the project is counted for each country.
    
    Args:
        all_projects_df (DataFrame): Combined DataFrame with all projects
    
    Returns:
        DataFrame: Statistics with country names and project counts
    """
    country_counts = {}
    
    # Check if required columns exist
    if 'recipient_country_m' not in all_projects_df.columns or 'recipient_country_count' not in all_projects_df.columns:
        return pd.DataFrame({'Country': [], 'Project Count': []})
    
    for _, row in all_projects_df.iterrows():
        recipient_country_m = row.get('recipient_country_m', '')
        recipient_country_count = row.get('recipient_country_count', 0)
        
        # Skip if no country data
        if not recipient_country_m or pd.isna(recipient_country_m):
            continue
        
        # Split by comma and strip whitespace
        countries = [country.strip() for country in str(recipient_country_m).split(',') if country.strip()]
        
        # Count this project for each country
        for country in countries:
            if country:
                country_counts[country] = country_counts.get(country, 0) + 1
    
    # Convert to DataFrame and sort by count (descending), then by country name
    if country_counts:
        stats_df = pd.DataFrame([
            {'Country': country, 'Project Count': count}
            for country, count in country_counts.items()
        ])
        stats_df = stats_df.sort_values(['Project Count', 'Country'], ascending=[False, True])
    else:
        stats_df = pd.DataFrame({'Country': [], 'Project Count': []})
    
    return stats_df

def generate_donor_statistics(all_projects_df):
    """
    Generate statistics of project counts by donors.
    When donor_count > 1, the project is counted for each donor.
    
    Args:
        all_projects_df (DataFrame): Combined DataFrame with all projects
    
    Returns:
        DataFrame: Statistics with donor names and project counts
    """
    donor_counts = {}
    
    # Check if required columns exist
    if 'donors' not in all_projects_df.columns or 'donor_count' not in all_projects_df.columns:
        return pd.DataFrame({'Donor': [], 'Project Count': []})
    
    for _, row in all_projects_df.iterrows():
        donors = row.get('donors', '')
        donor_count = row.get('donor_count', 0)
        
        # Skip if no donor data
        if not donors or pd.isna(donors):
            continue
        
        # Split by comma and strip whitespace
        donor_list = [donor.strip() for donor in str(donors).split(',') if donor.strip()]
        
        # Count this project for each donor
        for donor in donor_list:
            if donor:
                donor_counts[donor] = donor_counts.get(donor, 0) + 1
    
    # Convert to DataFrame and sort by count (descending), then by donor name
    if donor_counts:
        stats_df = pd.DataFrame([
            {'Donor': donor, 'Project Count': count}
            for donor, count in donor_counts.items()
        ])
        stats_df = stats_df.sort_values(['Project Count', 'Donor'], ascending=[False, True])
    else:
        stats_df = pd.DataFrame({'Donor': [], 'Project Count': []})
    
    return stats_df

def export_to_excel(manager_data_list, excel_filename):
    """
    Export multiple managers' projects to an Excel file with a single combined sheet.
    
    Args:
        manager_data_list (list): List of tuples (manager_id, DataFrame, manager_name, first_name)
        excel_filename (str): Name of the Excel file to create
    """
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        all_projects_list = []  # Collect all projects for the combined sheet
        
        for manager_id, df, manager_name, first_name in manager_data_list:
            if df is None or df.empty:
                continue
            
            # Add manager columns to the DataFrame
            df_with_manager = df.copy()
            df_with_manager.insert(0, 'manager_id', manager_id)
            df_with_manager.insert(1, 'manager', first_name)
            
            # Add to combined list
            all_projects_list.append(df_with_manager)
        
        # Create combined sheet with all projects
        if all_projects_list:
            all_projects_df = pd.concat(all_projects_list, ignore_index=True)
            
            # Reorder columns to put manager_id and manager first
            cols = all_projects_df.columns.tolist()
            if 'manager_id' in cols:
                cols.remove('manager_id')
                cols.insert(0, 'manager_id')
            if 'manager' in cols:
                cols.remove('manager')
                cols.insert(1, 'manager')
            all_projects_df = all_projects_df[cols]
            
            # Export to Excel (using default sheet name or 'Sheet1')
            all_projects_df.to_excel(writer, sheet_name='Sheet1', index=False)
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Sheet1']
            for idx, col in enumerate(all_projects_df.columns, 1):
                max_length = max(
                    all_projects_df[col].astype(str).map(len).max(),
                    len(str(col))
                )
                # Limit max width to 50 for readability
                adjusted_width = min(max_length + 2, 50)
                column_letter = get_column_letter(idx)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"  ✓ Sheet 'Sheet1' created with {len(all_projects_df)} rows (all projects combined)")
            
            # Generate and add recipient country statistics sheet
            stats_recipient_df = generate_country_statistics(all_projects_df)
            if not stats_recipient_df.empty:
                stats_recipient_df.to_excel(writer, sheet_name='statistics_recipient', index=False)
                
                # Auto-adjust column widths for statistics sheet
                worksheet_stats = writer.sheets['statistics_recipient']
                for idx, col in enumerate(stats_recipient_df.columns, 1):
                    max_length = max(
                        stats_recipient_df[col].astype(str).map(len).max(),
                        len(str(col))
                    )
                    adjusted_width = min(max_length + 2, 50)
                    column_letter = get_column_letter(idx)
                    worksheet_stats.column_dimensions[column_letter].width = adjusted_width
                
                print(f"  ✓ Sheet 'statistics_recipient' created with {len(stats_recipient_df)} countries")
            else:
                print(f"  ⚠ No recipient statistics generated (no country data found)")
            
            # Generate and add donor statistics sheet
            stats_donor_df = generate_donor_statistics(all_projects_df)
            if not stats_donor_df.empty:
                stats_donor_df.to_excel(writer, sheet_name='statistics_donor', index=False)
                
                # Auto-adjust column widths for statistics sheet
                worksheet_stats_donor = writer.sheets['statistics_donor']
                for idx, col in enumerate(stats_donor_df.columns, 1):
                    max_length = max(
                        stats_donor_df[col].astype(str).map(len).max(),
                        len(str(col))
                    )
                    adjusted_width = min(max_length + 2, 50)
                    column_letter = get_column_letter(idx)
                    worksheet_stats_donor.column_dimensions[column_letter].width = adjusted_width
                
                print(f"  ✓ Sheet 'statistics_donor' created with {len(stats_donor_df)} donors")
            else:
                print(f"  ⚠ No donor statistics generated (no donor data found)")

if __name__ == "__main__":
    # Read manager IDs from Excel file
    manager_id_file = 'manager id.xlsx'
    
    try:
        # Try to read manager IDs from the Excel file
        manager_df = pd.read_excel(manager_id_file, sheet_name='Sheet1')
        
        if 'manager_id' in manager_df.columns:
            # Extract manager IDs, filter out NaN/empty values, and convert to int
            manager_ids = manager_df['manager_id'].dropna().astype(int).unique().tolist()
            manager_ids = sorted(manager_ids)
            print(f"Loaded {len(manager_ids)} manager ID(s) from '{manager_id_file}'")
        else:
            print(f"Warning: 'manager_id' column not found in '{manager_id_file}'")
            print(f"Available columns: {list(manager_df.columns)}")
            print("Using default manager IDs.")
            manager_ids = [6820, 45014, 13416, 146624, 6316, 170987]
            
    except FileNotFoundError:
        print(f"Warning: File '{manager_id_file}' not found.")
        print("Using default manager IDs.")
        manager_ids = [6820, 45014, 13416, 146624, 6316, 170987]
    except Exception as e:
        print(f"Error reading '{manager_id_file}': {e}")
        print("Using default manager IDs.")
        manager_ids = [6820, 45014, 13416, 146624, 6316, 170987]
    
    if not manager_ids:
        print("Error: No manager IDs found. Exiting.")
        exit(1)
    
    print("=" * 60)
    print("UNIDO Compass Project Exporter")
    print("=" * 60)
    print(f"Processing {len(manager_ids)} manager(s)...")
    
    # Process each manager
    manager_data_list = []
    for manager_id in manager_ids:
        result = process_manager_projects(manager_id)
        if result[0] is not None:  # Check if DataFrame is not None
            df, manager_name, first_name = result
            manager_data_list.append((manager_id, df, manager_name, first_name))
    
    if not manager_data_list:
        print("\n✗ No data to export. Exiting.")
        exit(1)
    
    # Export to Excel
    if len(manager_ids) == 1:
        excel_filename = f"manager_{manager_ids[0]}_projects.xlsx"
    else:
        excel_filename = "managers_projects.xlsx"
    
    print(f"\n{'=' * 60}")
    print(f"Exporting to Excel: {excel_filename}")
    print(f"{'=' * 60}")
    
    export_to_excel(manager_data_list, excel_filename)
    
    print(f"\n{'=' * 60}")
    print(f"✓ Export completed: {excel_filename}")
    print(f"  Total managers processed: {len(manager_data_list)}")
    total_projects = sum(len(df) for _, df, _, _ in manager_data_list)
    print(f"  Total projects exported: {total_projects}")
    print(f"{'=' * 60}")

